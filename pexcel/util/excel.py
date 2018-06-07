#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Filename: excel.py
import openpyxl
import os
import re
import sys

' excel module '

__author__ = 'ituuz'

"""
本模块是完全处理excel所需要的所有接口
Excel 格式如下：
配表名字	template
描述		配表测试模版	
其他		作者:ituuz  日期：2018-04-24	
id			数值		说明
int			string		string
int			string	
id			value		des
KEY		
10000		test1		测试1
21001		test2		测试2
22001		test3		测试3

NOTE:
模版文件参考template目录下的t_template.xlsx文件
"""


def getFileName(file):
	fileName = os.path.basename(file)
	fileName = os.path.splitext(fileName)
	fileName = fileName[0]
	return fileName

def excelNameCheck(file, excelName):
	if excelName == None:
		excelName = getFileName(file)
	else:
		excelName = excelName.replace(" ", "")
		if excelName == "":
			excelName = getFileName(file)
		else:
			#判断首字母
			firstStr = excelName[0]
			if firstStr.isalpha():
				return excelName
			else:
				print("excel name it is not start with letters")
				sys.exit(0)
	return excelName

"""
将excel转换为数据列表
filename: excel文件路径
return：数据

数据格式：[
			[属性名1, 属性名2，属性名3],
			[数据类型1,数据类型2, 数据类型3],
			[数据1, 数据2, 数据3],
			[1, 0, 0],
			...
			...
			...
			[数据1, 数据2, 数据3]
		  ]
"""
def convertExcel2List(file, sheetname):
	print ("file name is : " + file)
	data = openpyxl.load_workbook(file)
	if(sheetname == ""):
		sheet = data.get_active_sheet()
	else:
		sheet = data.get_sheet_by_name(sheetname)

	listExcel = []
	excelName = sheet.cell(row=1, column=2).value
	listExcel = [excelNameCheck(file, excelName)]

	excelDes = sheet.cell(row=2, column=2).value
	excelOther = sheet.cell(row=3, column=2).value
	listExcel.append(excelDes)
	listExcel.append(excelOther)
	# print("listExcel", listExcel)

	# 属性名
	listProperty = []
	for rowProperty in sheet.iter_rows(min_row=4, max_col=5, max_row=4):
		for cell in rowProperty:
			listProperty.append(cell.value)
	listExcel.append(listProperty)
	# print("listExcel", listExcel)

	# 数据类型
	listNumClass = []
	# none列
	listNone = []
	col = 1
	for rowNumClass in sheet.iter_rows(min_row=6, max_col=5, max_row=6):
		for cell in rowNumClass:
			if cell.value == None :
				listNone.append(col)
			else:
				cell.value = cell.value.replace(" ", "")
				if cell.value == "" :
					listNone.append(col)
			listNumClass.append(cell.value)
			col = col + 1
	col = 1
	listExcel.append(listNumClass)
	# print("listNone==========", listNone, col)

	# 数据名
	listDataName = []
	for rowDataName in sheet.iter_rows(min_row=7, max_col=5, max_row=7):
		for cell in rowDataName:
			listDataName.append(cell.value)
	listExcel.append(listDataName)

	# key
	listKey = []
	for rowKey in sheet.iter_rows(min_row=8, max_col=5, max_row=8):
		for cell in rowKey:
			if cell.value == "KEY":
				listKey.append("1")
			else:
				listKey.append("0")
	listExcel.append(listKey)

	# 数据
	for row in sheet.iter_rows(min_row=9, max_col=5):
		listData = []
		for cell in row:
			# 空列不保存数据
			if col in listNone:
				col = col + 1
				continue
			listData.append(cell.value)
			col = col + 1
		col = 1
		listExcel.append(listData)

	return listExcel
	

def main():
	listExcel = convertExcel2List("../../template/t_template.xlsx", "")
	print(listExcel)
	# getFileName("../../template/t_template.xlsx")

if __name__=="__main__":
	main()


